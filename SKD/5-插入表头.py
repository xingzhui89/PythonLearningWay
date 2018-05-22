import tkinter.filedialog
from openpyxl import load_workbook
import openpyxl

#生成标准差的方程，输入一个列表
def stdDev(a):
    l = len(a)
    m = sum(a) / l
    d = 0
    for i in a:
        d += (i - m) ** 2
    return (d / (l - 1)) ** 0.5

#获取想要处理的Excel文件的文件名，返回一个列表
def getFile():
    result = []
    filenames = tkinter.filedialog.askopenfilenames()
    if len(filenames) != 0:
        for i in range(0, len(filenames)):
            string_filename = str(filenames[i])
            result.append(string_filename)
        return result
    else:
        print('您没有选择文件！')
        return


excelFiles = getFile()

#打开某一个Excel文件，并且对数据进行处理，返回的是一个列表
# 这个列表的每一个元素都是我们需要的数据组成的列表
def openExcel(stringname):
    result = []
    wb = load_workbook(filename=stringname)
    # print(wb.get_sheet_names())
    sheet = wb.get_sheet_by_name('Trenddata_1')
    rows = sheet.max_row
    # print(rows)
    for i in range(2, rows, 20):
        list = []
        PMP_Point_Name = sheet.cell(row=i, column=2).value
        # print(PMP_Point_Name)
        PMP_Direction = sheet.cell(row=i, column=3).value

        AV = []
        for j in range(7, 17):
            AV.append(sheet.cell(row=i, column=j).value)
        average = sum(AV) / len(AV)
        Min = min(AV)
        Max = max(AV)
        X = (Min + Max) / 2

        LT = sheet.cell(row=i + 1, column=7).value
        UT = sheet.cell(row=i + 2, column=7).value
        C = (LT + UT) / 2
        T = UT - LT
        Ca = (average - C) / T * 2
        # Ca=round(Ca,2)

        sixS = 6 * stdDev(AV)
        # sixS=round(sixS,2)
        # print(sixS)

        Cp = (UT - LT) / (sixS)
        # Cp=round(Cp,2)
        Cpk1 = Cp * (1 - abs(Ca))
        # Cpk1=round(Cpk1,2)
        Cpk = (Cpk1 if Cpk1 > 0 else 0)




        list.append(PMP_Point_Name)
        list.append(PMP_Direction)
        list.append(round(Cp, 2))
        list.append(round(Cpk, 2))
        list.append(Min)
        list.append(Max)
        list.append(LT)
        list.append(UT)
        list.append(round(X, 2))

        if (Cp<1) or (Cpk<1):
            result.append(list)
    return result

#保存列表内容到Excel中，通过弹窗获取输出文件名
def saveExcel(allPMP):
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = 'result'
    #表头
    sheet.cell(row=1, column=1).value = 'PMP_Point_Name'
    sheet.cell(row=1, column=2).value = 'PMP_Point_Name'
    sheet.cell(row=1, column=3).value = 'PMP_Direction'
    sheet.cell(row=1, column=4).value = 'Cp'
    sheet.cell(row=1, column=5).value = 'Cpk'
    sheet.cell(row=1, column=6).value = 'Min'
    sheet.cell(row=1, column=7).value = 'Max'
    sheet.cell(row=1, column=8).value = 'LT'
    sheet.cell(row=1, column=9).value = 'UT'
    sheet.cell(row=1, column=10).value = 'X'

    for row in range(2, len(allPMP) + 2):
        for col in range(1, 11):
            if col == 1:
                sheet.cell(row=row, column=col).value = allPMP[row - 2][col - 1]
            else:
                sheet.cell(row=row, column=col).value = allPMP[row - 2][col - 2]

    # for row in range(1, len(allPMP) + 1):
    #     sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)

    filename = tkinter.filedialog.asksaveasfilename()
    wb.save(filename=filename)


allPMP = []
for file in excelFiles:
    one_result = openExcel(file)
    for i in one_result:
        allPMP.append(i)

saveExcel(allPMP)
