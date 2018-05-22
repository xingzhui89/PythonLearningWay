import tkinter.filedialog
from openpyxl import load_workbook
import openpyxl


# 生成标准差的方程，输入一个列表
def stdDev(a):
    l = len(a)
    m = sum(a) / l
    d = 0
    for i in a:
        d += (i - m) ** 2
    return (d / (l - 1)) ** 0.5


# 获取想要处理的Excel文件的文件名，返回一个列表
def getFile():
    result = []
    filenames = tkinter.filedialog.askopenfilenames()
    if len(filenames) != 0:
        for i in range(0, len(filenames)):
            string_filename = str(filenames[i])
            result.append(string_filename)
        return result
    else:
        print('您没有选择文件！')
        return


excelFiles = getFile()


# 打开某一个Excel文件，并且对数据进行处理，返回的是一个列表
# 这个列表的每一个元素都是我们需要的数据组成的列表
def openExcel(stringname):
    result = []
    wb = load_workbook(filename=stringname)
    # print(wb.get_sheet_names())
    sheet = wb.get_sheet_by_name('Trenddata_1')
    rows = sheet.max_row
    # print(rows)
    print('打开文件：'+stringname)
    for i in range(2, rows, 20):
        list = []
        PMP_Point_Name = sheet.cell(row=i, column=2).value
        # print(PMP_Point_Name)
        PMP_Direction = sheet.cell(row=i, column=3).value

        AV = []
        if sheet.cell(row=i,column=7).value==None:
            continue

        for j in range(7, 17):
            AV.append(sheet.cell(row=i, column=j).value)

        average = sum(AV) / len(AV)
        Min = min(AV)
        Max = max(AV)
        X = (Min + Max) / 2

        LT = sheet.cell(row=i + 1, column=7).value
        UT = sheet.cell(row=i + 2, column=7).value
        C = (LT + UT) / 2
        T = UT - LT
        Ca = (average - C) / T * 2
        # Ca=round(Ca,2)

        sixS = 6 * stdDev(AV)
        # sixS=round(sixS,2)
        # 如果6six是0，
        if abs(sixS-0)<0.0000001:
            print(sixS, PMP_Point_Name)
            continue

        Cp = (UT - LT) / (sixS)
        # Cp=round(Cp,2)
        Cpk1 = Cp * (1 - abs(Ca))
        # Cpk1=round(Cpk1,2)
        Cpk = (Cpk1 if Cpk1 > 0 else 0)

        # --------------------添加上TMKS公差的计算方法----------------------------
        # 走了捷径，没有按照质保给的公式计算，而是通过直接平移或者扩大公差带的方法，反推CP和CPK以及PMPcheck合格的数值。
        TMKS_LT = 0
        TMKS_UT = 0
        NCp = 0
        if Cp < 1.0:
            for i in range(1, 10):
                TMKS_LT = LT - 0.1 * i
                TMKS_UT = UT + 0.1 * i
                NCp = (TMKS_UT - TMKS_LT) / (UT - LT) * Cp
                con1 = (TMKS_LT + TMKS_UT) / 2 + (TMKS_UT - TMKS_LT) * 0.75 / 2
                con2 = (TMKS_UT + TMKS_LT) / 2 - (TMKS_UT - TMKS_LT) * 0.75 / 2
                if NCp >= 1.0 and con1 >= Max and con2 <= Min:
                    break
            NNCPK = NCp * (1 - abs((X - (TMKS_UT + TMKS_LT) / 2) / (TMKS_UT - TMKS_LT) * 2))
            if NNCPK < 1.0:
                for i in range(-10, 10):
                    TMKS_LT += i * 0.1
                    TMKS_UT += i * 0.1
                    NNNCpk = Cp * (1 - abs((X - (TMKS_UT + TMKS_LT) / 2) / (TMKS_UT - TMKS_LT) * 2))
                    con1 = (TMKS_LT + TMKS_UT) / 2 + (TMKS_UT - TMKS_LT) * 0.75 / 2
                    con2 = (TMKS_UT + TMKS_LT) / 2 - (TMKS_UT - TMKS_LT) * 0.75 / 2
                    if NNNCpk >= 1.0 and con1 >= Max and con2 <= Min:
                        break
        elif Cpk < 1.0:
            for i in range(-10, 10):
                TMKS_LT = LT + i * 0.1
                TMKS_UT = UT + i * 0.1
                NCpk = Cp * (1 - abs((X - (TMKS_UT + TMKS_LT) / 2) / (TMKS_UT - TMKS_LT) * 2))
                con1 = (TMKS_LT + TMKS_UT) / 2 + (TMKS_UT - TMKS_LT) * 0.75 / 2
                con2 = (TMKS_UT + TMKS_LT) / 2 - (TMKS_UT - TMKS_LT) * 0.75 / 2
                if NCpk >= 1.0 and con1 >= Max and con2 <= Min:
                    break

        # -------------------------------------------------------

        list.append(PMP_Point_Name)
        list.append(PMP_Direction)
        list.append(round(Cp, 2))
        list.append(round(Cpk, 2))
        list.append(Min)
        list.append(Max)
        list.append(LT)
        list.append(UT)
        list.append(round(X, 2))
        list.append(round(TMKS_LT, 1))
        list.append(round(TMKS_UT, 1))

        if (Cp < 1) or (Cpk < 1):
            result.append(list)
    return result


# 保存列表内容到Excel中，通过弹窗获取输出文件名
def saveExcel(allPMP):
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = 'result'
    # 表头
    sheet.cell(row=1, column=1).value = 'PMP_Point_Name'
    sheet.cell(row=1, column=2).value = 'PMP_Point_Name'
    sheet.cell(row=1, column=3).value = 'PMP_Direction'
    sheet.cell(row=1, column=4).value = 'Cp'
    sheet.cell(row=1, column=5).value = 'Cpk'
    sheet.cell(row=1, column=6).value = 'Min'
    sheet.cell(row=1, column=7).value = 'Max'
    sheet.cell(row=1, column=8).value = 'LT'
    sheet.cell(row=1, column=9).value = 'UT'
    sheet.cell(row=1, column=10).value = 'X'
    sheet.cell(row=1, column=11).value = 'TMKS_LT'
    sheet.cell(row=1, column=12).value = 'TMKS_UT'

    for row in range(2, len(allPMP) + 2):
        for col in range(1, 13):
            if col == 1:
                sheet.cell(row=row, column=col).value = allPMP[row - 2][col - 1]
            else:
                sheet.cell(row=row, column=col).value = allPMP[row - 2][col - 2]

    # for row in range(1, len(allPMP) + 1):
    #     sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)

    filename = tkinter.filedialog.asksaveasfilename()
    filename=filename+'.xlsx'
    wb.save(filename=filename)


allPMP = []
for file in excelFiles:
    one_result = openExcel(file)
    for i in one_result:
        allPMP.append(i)

saveExcel(allPMP)
