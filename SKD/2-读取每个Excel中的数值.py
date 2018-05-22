'''
http://blog.sina.com.cn/s/blog_03f96e310102x0dg.html
https://www.cnblogs.com/joyeecheung/p/3842002.html
以上为excel操作的参考
'''
import tkinter.filedialog
from openpyxl import load_workbook

def stdDev(a):
    l=len(a)
    m=sum(a)/l
    d=0
    for i in a:
        d+=(i-m)**2
    return (d/(l-1))**0.5

def getFile():
    result=[]
    filenames = tkinter.filedialog.askopenfilenames()
    if len(filenames) != 0:
        for i in range(0, len(filenames)):
            string_filename = str(filenames[i])
            result.append(string_filename)
        return result
    else:
        print('您没有选择文件！')
        return

excelFiles=getFile()

def openExcel(stringname):
    result=[]
    wb=load_workbook(filename=stringname)
    # print(wb.get_sheet_names())
    sheet=wb.get_sheet_by_name('Trenddata_1')
    rows=sheet.max_row
    # print(rows)
    for i in range(2,rows,20):
        list=[]
        PMP_Point_Name=sheet.cell(row=i,column=2).value
        #print(PMP_Point_Name)
        PMP_Direction=sheet.cell(row=i,column=3).value

        AV=[]
        for j in range(7,17):
            AV.append(sheet.cell(row=i,column=j).value)
        average=sum(AV)/len(AV)
        Min=min(AV)
        Max=max(AV)

        LT=sheet.cell(row=i+1,column=7).value
        UT=sheet.cell(row=i+2,column=7).value
        C=(LT+UT)/2
        T=UT-LT
        Ca=(average-C)/T*2
        # Ca=round(Ca,2)

        sixS=6*stdDev(AV)
        # sixS=round(sixS,2)
        #print(sixS)

        Cp=(UT-LT)/(sixS)
        # Cp=round(Cp,2)
        Cpk1=Cp*(1-abs(Ca))
        # Cpk1=round(Cpk1,2)
        Cpk=(Cpk1 if Cpk1>0 else 0)

        list.append(PMP_Point_Name)
        list.append(PMP_Direction)
        list.append(round(Cp,2))
        list.append(round(Cpk,2))
        list.append(Min)
        list.append(Max)
        list.append(LT)
        list.append(UT)

        result.append(list)
    return result

for file in excelFiles:
    allPMP=openExcel(file)
    for i in allPMP:
        print(i)